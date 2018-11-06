from bs4 import BeautifulSoup
import time
import requests
import time
from openpyxl import Workbook
import numpy as np
from urllib import parse, error
import urllib
import random

# Some User Agents
heads = [
    {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.77 Safari/537.36'},
    {'User-Agent':'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11'},
    {'User-Agent':'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0'}
]
# book_tag 是书的类型
def book_spider(book_tag):
    page_num = 0
    book_list = []
    try_times = 0
    while(1):
        # url = 'http://www.douban.com/tag/%E5%B0%8F%E8%AF%B4/book?start=0' For test
        #url = 'https://www.douban.com/tag/%E5%B0%8F%E8%AF%B4/' + '/book?start=' + str(page_num * 15)
        url = 'https://www.douban.com/tag/'+parse.quote(book_tag) + '/book?start=' + str(page_num * 15)
        time.sleep(3)
        # time.sleep(random.randint(5))
        try:
            req = requests.get(url=url, headers=heads[random.randint(0, 2)])
            source_code = req.text
            plain_text = str(source_code)
        except error.HTTPError as e:
            print(e)
            continue
        except error.URLError as e:
            print(e)
            continue

        soup = BeautifulSoup(plain_text, 'html.parser')
        list_soup = soup.find('div', {'class': "mod book-list"})
        try_times += 1

        # 加入程序停止判断
        if list_soup == None and try_times < 200:
            continue
        elif list_soup == None or len(list_soup) <= 1:
            break
        for book_info in list_soup('dd'):
            title = book_info.find('a', {'class': 'title'}).string.strip()
            desc = book_info.find('div', {'class': 'desc'}).string.strip()
            # 去掉str中的'/'
            desc_list = desc.split('/')
            book_url = book_info.find('a', {'class': 'title'}).get('href')
            # print(book_url)  For test
            # 获取数据
            try:
                author_info = '作者译者: ' + '/'.join(desc_list[0:-3])
            except:
                author_info = '作者/译者： 暂无'
            try:
                pub_info = '出版信息：' + '/'.join(desc_list[-3:0])
            except:
                pub_info = '出版信息：暂无'
            # 获取书的评分
            try:
                rating = book_info.find('div', {'class': 'rating_nums'}).findAll('span')[1].string.strip()
            except:
                rating = '0.0'
            # 通过book_url内的网址,进入书的详情页面获取信息
            try:
                people_num = get_people_num(book_url)
            except:
                people_num = '0'
            # 把获取到的数据放入 book_list中
            # 标题,评分,评价人数,作者,出版信息
            book_list.append([title, rating, people_num, author_info, pub_info])
            try_times = 0
        print(book_list)
        page_num += 1

        print('Download Information From Page %d' % page_num)
    return book_list
# 进入书详情页面获取对书籍的评价人数
def get_people_num(url):
    # https://book.douban.com/subject/6518605/?from=tag_allc  # for test
    global plain_text_1
    try:
        req = requests.get(url, headers=heads[random.randint(0, 2)])
        source_code = req.text
        plain_text_1 = str(source_code)
    except error.HTTPError as e:
        print(e)
    except error.URLError as e:
        print(e)
    soup = BeautifulSoup(plain_text_1, 'html.parser')    # 解释器： html.parse
    people_num = soup.find('div', {'class': 'rating_sum'}).findAll('span')[1].string.strip()
    return people_num

def do_spider(book_tag_lists):
    book_lists = []
    for book_tag in book_tag_lists:
        book_list = book_spider(book_tag)
        print(book_list)
        book_list = sorted(book_list, key=lambda x: x[1], reverse=True)
        book_lists.append(book_list)
    return book_lists

# 把获取的数据存入excal 中
def print_book_lists_excel(book_lists,book_tag_lists):
    wb = Workbook()
    ws = []
    for i in range(len(book_tag_lists)):
        ws.append(wb.create_sheet(title=book_tag_lists[i].decode())) #utf8->unicode
    for i in range(len(book_tag_lists)):
        ws[i].append(['序号', '书名', '评分', '评价人数', '作者', '出版社'])
        count=1
        for bl in book_lists[i]:
            ws[i].append([count, bl[0], float(bl[1]), int(bl[2]), bl[3], bl[4]])
            count += 1
    save_path = 'book_list'
    for i in range(len(book_tag_lists)):
        save_path += ('-'+book_tag_lists[i].decode())
    save_path += '.xlsx'
    wb.save(save_path)
if __name__ == '__main__':
    #book_tag_lists = ['个人管理', '时间管理', '投资', '文化', '宗教']
    book_tag_lists = ['个人管理']
    do_spider(book_tag_lists)
    # print_book_lists_excel ( book_lists, book_tag_lists )